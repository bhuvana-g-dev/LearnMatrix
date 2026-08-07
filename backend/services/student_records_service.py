"""
services/student_records_service.py

Admin-only. Joins three independent per-uid Firestore collections
(services/assessment_repository.py, services/roadmap_repository.py,
services/activity_repository.py) plus each uid's email from Firebase
Auth itself (firebase_admin.auth — already initialized by
firebase/firebase_config.py, no new credential needed) into ONE
real, non-fake picture of "what has this student actually done":
which diagnostic they took, what every skill scored, and how far their
roadmap has progressed.

Two shapes come out of this:
  - get_student_summaries(): one row per student — for the on-screen
    admin table.
  - get_quiz_attempts(): one row per (student, assessed skill) — the
    actual "what quizzes did they attend" detail, since each skill's
    diagnostic section IS a quiz attempt in this app's model (a fixed
    set of easy/medium/hard questions, scored, leveled).

build_export_workbook() turns both into a real .xlsx file via
openpyxl (already a dependency — see backend/requirements.txt) — no
new package needed, and this never runs client-side, so there's no
risk of a huge roster freezing anyone's browser.

Nothing here is invented: every column traces to an actual Firestore
field or a value computed directly from one (e.g. completionPercent is
read straight off the saved roadmap doc, not recomputed).
"""

from io import BytesIO

from firebase_admin import auth as firebase_auth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from firebase.firebase_config import get_firestore_client
from services.assessment_repository import list_all_assessment_results
from services.roadmap_repository import list_all_roadmaps
from services.activity_repository import get_activity_dates


def _email_for_uid(uid: str) -> str:
    """Best-effort — a uid can outlive its Auth record (deleted account,
    emulator data, etc.), so this never raises; '—' just means "not
    resolvable right now", not an error in the export itself."""
    try:
        return firebase_auth.get_user(uid).email or "—"
    except Exception:  # noqa: BLE001 — any Auth lookup failure degrades the same way
        return "—"


def get_student_summaries() -> list[dict]:
    """One row per student who has completed at least one assessment —
    students who never started are simply absent, not an error."""
    db = get_firestore_client()
    assessments = {a["uid"]: a for a in list_all_assessment_results(db) if a.get("uid")}
    roadmaps = {r["uid"]: r for r in list_all_roadmaps(db) if r.get("uid")}

    summaries = []
    for uid, assessment in assessments.items():
        roadmap = roadmaps.get(uid)
        evaluation = assessment.get("evaluation") or {}
        overall = evaluation.get("overall") or {}
        active_days = len(get_activity_dates(db, uid))

        summaries.append({
            "uid": uid,
            "email": _email_for_uid(uid),
            "role": assessment.get("role", ""),
            "skillsAssessed": len(evaluation.get("skills") or []),
            "overallScorePercent": overall.get("scorePercent"),
            "submittedAt": assessment.get("submittedAt"),
            "roadmapTotalSkills": roadmap.get("totalSkills") if roadmap else None,
            "roadmapMasteredCount": roadmap.get("masteredCount") if roadmap else None,
            "roadmapCompletionPercent": roadmap.get("courseCompletionPercent") if roadmap else None,
            "activeDays": active_days,
        })

    summaries.sort(key=lambda s: s["email"])
    return summaries


def get_quiz_attempts() -> list[dict]:
    """One row per (student, assessed skill) — the per-skill diagnostic
    quiz result, which is the actual gradeable "attempt" in this app's
    model (services/evaluation_service.py's SkillResult)."""
    db = get_firestore_client()
    assessments = list_all_assessment_results(db)

    attempts = []
    for a in assessments:
        uid = a.get("uid")
        if not uid:
            continue
        email = _email_for_uid(uid)
        role = a.get("role", "")
        submitted_at = a.get("submittedAt")
        for skill_result in (a.get("evaluation") or {}).get("skills") or []:
            breakdown = skill_result.get("breakdown") or {}
            attempts.append({
                "uid": uid,
                "email": email,
                "role": role,
                "skill": skill_result.get("skill", ""),
                "level": skill_result.get("level", ""),
                "scorePercent": skill_result.get("scorePercent"),
                "easyCorrect": breakdown.get("easy_correct"),
                "easyTotal": breakdown.get("easy_total"),
                "mediumCorrect": breakdown.get("medium_correct"),
                "mediumTotal": breakdown.get("medium_total"),
                "hardCorrect": breakdown.get("hard_correct"),
                "hardTotal": breakdown.get("hard_total"),
                "submittedAt": submitted_at,
            })

    attempts.sort(key=lambda r: (r["email"], r["skill"]))
    return attempts


_HEADER_FILL = PatternFill(start_color="7C6FE0", end_color="7C6FE0", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers: list[tuple[str, str]], rows: list[dict]) -> None:
    """headers: [(column_label, row_dict_key), ...] — keeps the label
    and the source key next to each other so a column can't silently
    end up reading the wrong field."""
    for col_idx, (label, _key) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(headers, start=1):
            value = row.get(key)
            # Firestore timestamps aren't natively Excel-writable —
            # openpyxl handles real datetimes fine, but a raw Firestore
            # Timestamp object needs isoformat() first.
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, (label, _key) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 2, 14)


def build_export_workbook() -> BytesIO:
    """Two-sheet .xlsx: Student Summary + Quiz Attempts (per skill).
    Returns an in-memory BytesIO ready to stream as a file response —
    nothing is written to disk on the server."""
    summaries = get_student_summaries()
    attempts = get_quiz_attempts()

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Student Summary"
    _write_sheet(
        summary_ws,
        [
            ("Email", "email"), ("UID", "uid"), ("Role", "role"),
            ("Skills Assessed", "skillsAssessed"), ("Overall Score %", "overallScorePercent"),
            ("Assessment Date", "submittedAt"), ("Roadmap Total Skills", "roadmapTotalSkills"),
            ("Roadmap Mastered", "roadmapMasteredCount"), ("Roadmap Completion %", "roadmapCompletionPercent"),
            ("Active Days", "activeDays"),
        ],
        summaries,
    )

    attempts_ws = wb.create_sheet("Quiz Attempts")
    _write_sheet(
        attempts_ws,
        [
            ("Email", "email"), ("UID", "uid"), ("Role", "role"), ("Skill", "skill"),
            ("Level", "level"), ("Score %", "scorePercent"),
            ("Easy Correct", "easyCorrect"), ("Easy Total", "easyTotal"),
            ("Medium Correct", "mediumCorrect"), ("Medium Total", "mediumTotal"),
            ("Hard Correct", "hardCorrect"), ("Hard Total", "hardTotal"),
            ("Assessment Date", "submittedAt"),
        ],
        attempts,
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
